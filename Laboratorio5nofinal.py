from virus_total_apis import PublicApi
from datetime import date
from openpyxl import Workbook
import time

#Creamos El Excel
libro = Workbook()
hoja = libro.active

#Nombramos La Hoja
hoja1 = libro.create_sheet("Laboratorio5")
libro.remove_sheet(libro.get_sheet_by_name("Sheet"))


#Creamos las variables de las celdas
numlinea = 2
numlinea2 = 2
numlinea3 = 2
numlinea4 = 2
numlinea5 = 2
columna = 1
fila = 1

a1 = hoja1['A1']
a1.value = "URL"
b1 = hoja1['B1']
b1.value = "Fecha de análisis"
c1 = hoja1['C1']
c1.value = "Total de análisis"
d1 = hoja1['D1']
d1.value = "Análisis positivos"
e1 = hoja1['E1']
e1.value = "Clasificación"

#Leemos el archivo txt
f = open("Url.txt","r")
lines = f.readlines()
f.close()

#introducimos la API KEY
API_KEY = "558336c3a4949fe01bb7830f43c3f12af88414012661b3abfb2c80693c0ae2a4"
api = PublicApi(API_KEY)

#Hacer un for que acabe cuando termine de leer las url
for url in lines:

    borrar = "\n"
    url = url.replace(borrar, "")
    print(url)
    hoja1.cell(numlinea, 1, url)

    time.sleep(15)
    response = api.get_url_report(url)

    numlinea += 1
    #d3 fecha de analisi
    today = date.today()
    d3 = today.strftime("%d/%m/%Y")
    print("Dia de hoy", d3)
    hoja1.cell(numlinea2, 2, d3)
    numlinea2 += 1
    if response["response_code"] == 200:
           #Analisis positivos
           # con un if se hace la clasificacion
        print(response["results"]["positives"])
        hoja1.cell(numlinea3, 4, (response["results"]["positives"]))
        numlinea3 += 1
        if (response["results"]["positives"]) <= 3:
            print("Clasificacion baja")
            hoja1.cell(numlinea4, 5, "Baja")
            numlinea4 += 1
        elif (response["results"]["positives"]) < 10:
            hoja1.cell(numlinea4, 5, "Media")
            numlinea4 += 1
        else:
            print("Clasificacion alta")
            hoja1.cell(numlinea4, 5, "Alta")
            numlinea4 += 1
    #total de analisis
        print(response["results"]["total"])
        hoja1.cell(numlinea5, 3, (response["results"]["total"]))
        numlinea5 += 1
    else:
        print("No ha podido obtenerse el análisis del archivo.")
        

libro.save("reporte_analizador_urls.xlsx")
